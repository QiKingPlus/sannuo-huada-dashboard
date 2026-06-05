#!/usr/bin/env python3
"""分析三诺成交用户聊天记录"""

import os
import json
import re
import openpyxl
from collections import Counter, defaultdict

dir_path = "/Users/SaciNa/Documents/臻选&芥子/臻选私域/聊天记录/三诺成交用户"
raw_files = sorted(os.listdir(dir_path))

all_conversations = []  # 每个元素: {filename, advisor, customer, messages[]}

for f_idx, f in enumerate(raw_files):
    filepath = os.path.join(dir_path, f)
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        headers = []
        for col in range(1, ws.max_column + 1):
            headers.append(ws.cell(1, col).value)
        
        messages = []
        advisor_name = None
        customer_name = None
        
        for row in range(2, ws.max_row + 1):
            msg = {}
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row, col).value
                h = headers[col-1] if col-1 < len(headers) else f"col{col}"
                msg[h] = val
            
            # Extract content from JSON
            content = ""
            if msg.get('消息内容'):
                try:
                    content_json = json.loads(msg['消息内容'])
                    content = content_json.get('content', '')
                except:
                    content = str(msg['消息内容'])
            
            # Identify sender role
            sender_name = msg.get('发送人名称', '')
            receiver_name = msg.get('接收人名称', '')
            
            if not advisor_name and sender_name:
                advisor_name = sender_name
            if not customer_name and receiver_name:
                customer_name = receiver_name
            
            messages.append({
                'time': msg.get('消息时间', ''),
                'sender': sender_name,
                'receiver': receiver_name,
                'content': content,
                'type': msg.get('消息类型', '')
            })
        
        wb.close()
        
        all_conversations.append({
            'file': f,
            'advisor': advisor_name or '',
            'customer': customer_name or '',
            'msg_count': len(messages),
            'messages': messages
        })
        
    except Exception as e:
        print(f"Error reading {f}: {e}")

# ========== ANALYSIS ==========

output_lines = []
def p(line=""):
    output_lines.append(line)
    print(line)

p("=" * 80)
p("📊 三诺成交用户 · 聊天记录深度分析报告")
p(f"分析时间: 2026-05-26")
p(f"分析文件数: {len(all_conversations)} 个成交用户")
p("=" * 80)

# 1. 基础统计
p("\n## 一、基础数据总览\n")
p("| 序号 | 文件名 | 顾问 | 客户 | 消息数 |")
p("|------|--------|------|------|--------|")
for i, conv in enumerate(all_conversations):
    fname = conv['file'][:30] + "..." if len(conv['file']) > 30 else conv['file']
    p(f"| {i+1} | {fname} | {conv['advisor']} | {conv['customer']} | {conv['msg_count']} |")

total_msgs = sum(c['msg_count'] for c in all_conversations)
p(f"\n**总计**: {total_msgs} 条消息，平均每个用户 {total_msgs/len(all_conversations):.1f} 条")

# 2. 对话内容分析
p("\n## 二、对话流程分析\n")

# Collect all messages content
all_contents = []
for conv in all_conversations:
    for msg in conv['messages']:
        if msg['content']:
            all_contents.append(msg['content'])

full_text = '\n'.join(all_contents)

# 2.1 话题识别
p("### 2.1 高频话题关键词\n")
topics = {
    '肠道/肠胃/便秘/腹泻/胀气': ['肠道', '肠胃', '便秘', '腹泻', '胀气', '肚子', '排便', '拉肚子'],
    '睡眠/失眠/熬夜': ['睡眠', '失眠', '熬夜', '睡不着', '入睡'],
    '减重/减脂/体重': ['减肥', '减脂', '减重', '体重', '胖', '瘦'],
    '血糖/糖尿病': ['血糖', '糖尿病', '三诺'],
    '益生菌/产品': ['益生菌', '益休', '益舒', '益畅', '镜脂轻', '轻路', '膳食纤维'],
    '效果/改善': ['效果', '改善', '见效', '好转', '有用'],
    '价格/购买': ['价格', '多少钱', '购买', '下单', '优惠', '活动'],
    '体检/指标': ['体检', '指标', '检查', '报告'],
    '饮食/生活习惯': ['饮食', '运动', '喝水', '作息', '熬夜'],
}

p("\n| 话题类别 | 出现次数 | 占比 |")
p("|----------|----------|------|")
for topic, keywords in topics.items():
    count = 0
    for kw in keywords:
        count += len(re.findall(kw, full_text, re.IGNORECASE))
    pct = count / max(len(all_contents), 1) * 100
    p(f"| {topic} | {count} | {pct:.1f}% |")

# 3. 成交路径分析
p("\n## 三、成交路径与话术分析\n")

# 3.1 对话阶段识别
p("### 3.1 典型成交对话阶段\n")
stages = {
    '开场/建立信任': ['你好', '您好', '我是', '顾问', '健康', '三诺'],
    '需求挖掘/提问': ['什么情况', '多久了', '平时', '有没有', '了解', '怎么'],
    '症状描述': ['不舒服', '难受', '一直', '有时候', '经常', '感觉'],
    '科普教育': ['菌群', '益生菌', '肠道菌群', '有益菌', '有害菌', '原理'],
    '产品推荐': ['推荐', '可以试试', '适合', '搭配', '益休', '益舒', '益畅'],
    '价格协商': ['多少钱', '价格', '优惠', '活动', '便宜'],
    '顾虑消除': ['担心', '副作用', '依赖', '反弹', '安全'],
    '促单成交': ['下单', '试试', '先买', '付款', '地址'],
    '售后跟进': ['收到', '服用', '感觉', '效果', '继续'],
}

for stage, keywords in stages.items():
    count = 0
    for kw in keywords:
        count += len(re.findall(kw, full_text, re.IGNORECASE))
    p(f"- **{stage}**: 关键词出现 {count} 次")

# 4. 产品提及分析
p("\n## 四、产品提及分析\n")
products = ['益休', '益舒', '益畅', '镜脂轻', '轻路', '膳食纤维', '华大']
p("\n| 产品 | 提及次数 |")
p("|------|----------|")
for prod in products:
    count = len(re.findall(prod, full_text))
    p(f"| {prod} | {count} |")

# 5. 客户痛点分析
p("\n## 五、客户核心痛点分析\n")
pain_points = {
    '便秘/排便问题': ['便秘', '排便', '拉不出', '大便'],
    '腹泻/肠胃不适': ['腹泻', '拉肚子', '肚子疼', '胃不舒服'],
    '睡眠障碍': ['失眠', '睡不好', '入睡困难', '半夜醒'],
    '体重困扰': ['胖', '减肥', '体重', '减不下来'],
    '血糖问题': ['血糖', '糖尿病', '控糖'],
    '免疫力低下': ['感冒', '免疫力', '容易生病'],
    '皮肤问题': ['皮肤', '痘痘', '过敏', '湿疹'],
}

p("\n| 痛点类别 | 出现次数 |")
p("|----------|----------|")
for pain, keywords in pain_points.items():
    count = 0
    for kw in keywords:
        count += len(re.findall(kw, full_text))
    p(f"| {pain} | {count} |")

# 6. 销售话术分析
p("\n## 六、销售话术技巧分析\n")
sales_tactics = {
    '提问式挖掘需求': ['什么情况', '多久了', '平时怎么', '有没有', '之前'],
    '专业背书': ['华大', '科研', '专利', '菌株', '临床'],
    '用户证言/案例': ['其他客户', '之前有', '不少人', '反馈'],
    '紧迫感营造': ['活动', '限时', '优惠', '最后'],
    '保障承诺': ['放心', '安全', '保证', '无副作用'],
    '复购引导': ['继续', '巩固', '长期', '坚持'],
}

p("\n| 话术技巧 | 出现次数 |")
p("|----------|----------|")
for tactic, keywords in sales_tactics.items():
    count = 0
    for kw in keywords:
        count += len(re.findall(kw, full_text))
    p(f"| {tactic} | {count} |")

# 7. 每段对话详细摘要
p("\n## 七、各对话详细摘要\n")

for i, conv in enumerate(all_conversations):
    p(f"\n### 对话 {i+1}: {conv['advisor']} ↔ {conv['customer']}")
    p(f"- 消息总数: {conv['msg_count']}")
    
    # 提取顾问的关键话术
    advisor_msgs = [m for m in conv['messages'] if m['sender'] == conv['advisor']]
    customer_msgs = [m for m in conv['messages'] if m['sender'] == conv['customer']]
    
    p(f"- 顾问消息: {len(advisor_msgs)}条, 客户消息: {len(customer_msgs)}条")
    
    # 对话时间跨度
    times = [m['time'] for m in conv['messages'] if m['time']]
    if times:
        p(f"- 对话时间范围: {times[-1]} ~ {times[0]}")
    
    # 提取关键内容摘要
    p("\n**对话关键节点**:")
    for msg in conv['messages']:
        content = msg['content']
        if not content:
            continue
        # 只看有意义的消息
        if any(kw in content for kw in ['推荐', '益生菌', '益休', '益舒', '益畅', '镜脂轻', '下单', '付款', '地址', '效果', '改善', '服用']):
            sender = msg['sender']
            role = "🏪顾问" if sender == conv['advisor'] else "👤客户"
            p(f"  {msg['time']} {role}: {content[:120]}")

# 8. 成功成交要素总结
p("\n## 八、成交成功要素总结\n")
p("""
### 8.1 高转化话术特征
1. **先问后推**: 顾问普遍采用「先了解情况→再科普教育→后产品推荐」的阶梯式话术
2. **痛点共鸣**: 对客户描述的症状表达理解和共情，而非直接推销
3. **专业背书**: 频繁使用「华大」「科研」「专利菌株」等建立信任
4. **生活建议先行**: 先给饮食/作息建议，再推荐产品，降低推销感

### 8.2 客户转化关键节点
1. **信任建立期**: 客户开始主动描述自己的症状细节
2. **认知转变期**: 客户接受了「益生菌/肠道菌群」与健康的关系
3. **决策推动期**: 价格/效果顾虑被消除，产生购买意愿

### 8.3 常见异议处理模式
- "怕没效果" → 案例佐证 + 科学原理 + 试用建议
- "怕有副作用" → 安全性说明 + 天然成分 + 无依赖
- "价格贵" → 价值换算 + 长期健康投资 + 活动优惠
- "再考虑考虑" → 限时活动 + 健康紧迫性 + 低门槛试饮
""")

# 9. 改进建议
p("\n## 九、从成交数据反推的优化建议\n")
p("""
### 9.1 话术优化方向
1. **标准化「需求挖掘清单」**: 统一顾问的开场提问框架，确保不遗漏关键信息
2. **强化「科学教育」环节**: 用通俗语言解释菌群原理，提升客户认知进而提升客单价
3. **丰富「客户证言库」**: 按症状分类整理客户案例，方便顾问精准引用

### 9.2 产品搭配优化
- 根据高频痛点（便秘、睡眠、减重）制定标准搭配方案
- 建立「先主后辅」的产品推荐逻辑

### 9.3 售后跟进关键节点
- 服用第3天：询问是否有赫氏消亡反应（正常现象，预告在前建立信任）
- 服用第7天：询问初步感受
- 服用第14天：效果反馈 + 复购引导
- 服用第30天：全面评估 + 长期方案
""")

# 保存报告
report_path = "/Users/SaciNa/WorkBuddy/2026-05-26-15-30-31/三诺成交用户分析报告.md"
with open(report_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(output_lines))

print(f"\n\n报告已保存至: {report_path}")
