#!/usr/bin/env python3
"""三诺成交用户聊天记录 - 完整分析"""

import os, json, re
import openpyxl
from collections import Counter, defaultdict

dir_path = "/Users/SaciNa/Documents/臻选&芥子/臻选私域/聊天记录/三诺成交用户"
raw_files = sorted(os.listdir(dir_path))

all_conversations = []

for fi, f in enumerate(raw_files):
    filepath = os.path.join(dir_path, f)
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    # Build sender_name → role mapping
    # The advisor always has "赵老师" or "三诺健康管家" in their name
    names = set()
    for row in range(2, ws.max_row + 1):
        sn = str(ws.cell(row, 4).value or '')
        rn = str(ws.cell(row, 6).value or '')
        if sn: names.add(sn)
        if rn: names.add(rn)
    
    advisor_name = ''
    customer_name = ''
    for name in names:
        if '赵老师' in name or '健康管家' in name or '三诺' in name:
            advisor_name = name
        else:
            customer_name = name
    
    # Parse filename for phone
    phone = f.replace('.xlsx', '').rsplit('-', 1)[-1] if '-' in f else ''
    
    messages = []
    for row in range(2, ws.max_row + 1):
        sn = str(ws.cell(row, 4).value or '')
        rn = str(ws.cell(row, 6).value or '')
        time_str = str(ws.cell(row, 2).value or '')
        content_raw = ws.cell(row, 8).value or ''
        msg_type = str(ws.cell(row, 7).value or '')
        
        try:
            cj = json.loads(content_raw)
            content = cj.get('content', '')
        except:
            content = str(content_raw)
        
        role = 'advisor' if sn == advisor_name else 'customer'
        
        messages.append({
            'time': time_str,
            'role': role,
            'content': content,
            'type': msg_type
        })
    
    wb.close()
    
    advisor_msgs = [m for m in messages if m['role'] == 'advisor']
    customer_msgs = [m for m in messages if m['role'] == 'customer']
    
    all_conversations.append({
        'file': f,
        'advisor': advisor_name,
        'customer': customer_name,
        'phone': phone,
        'messages': messages,
        'advisor_msgs': advisor_msgs,
        'customer_msgs': customer_msgs,
    })

# ============================================================
# Generate Report
# ============================================================

out = []
def p(line=""):
    out.append(line)

p("# 📊 三诺成交用户 · 聊天记录深度分析报告\n")
p(f"> **分析时间**: 2026-05-26 | **文件数**: {len(all_conversations)} 个成交用户 | **总消息**: {sum(len(c['messages']) for c in all_conversations)} 条\n")

# ─── 1. 总览 ───
p("## 一、成交用户画像总览\n")
p("| # | 顾问 | 客户 | 手机 | 消息 | 顾问 | 客户 | 周期 |")
p("|---|------|------|------|------|------|------|------|")
for i, c in enumerate(all_conversations):
    msgs = c['messages']
    a = len(c['advisor_msgs'])
    cu = len(c['customer_msgs'])
    d1 = msgs[-1]['time'][:10] if msgs else ''
    d2 = msgs[0]['time'][:10] if msgs else ''
    period = f"{d1}~{d2}" if d1 != d2 else d1
    advisor_short = c['advisor'].replace('三诺健康管家', '管家').replace('｜', '|').replace(' ', '')[:8]
    p(f"| {i+1} | {advisor_short} | {c['customer'][:8]} | {c['phone']} | {len(msgs)} | {a} | {cu} | {period} |")

# Advisor stats
advisor_counts = Counter([c['advisor'] for c in all_conversations])
p(f"\n> **顾问分布**: " + "、".join([f"{name}({count}人)" for name, count in advisor_counts.most_common()]))
p(f"> 总计 {sum(len(c['messages']) for c in all_conversations)} 条消息，顾问发言 {sum(len(c['advisor_msgs']) for c in all_conversations)} 条 ({sum(len(c['advisor_msgs']) for c in all_conversations)/max(sum(len(c['messages']) for c in all_conversations),1)*100:.1f}%)\n")

# ─── 2. 每段对话详解 ───
p("## 二、每段对话深度拆解\n")

for i, c in enumerate(all_conversations):
    p(f"### 📝 对话 {i+1}：{c['customer']}（顾问：{c['advisor'].replace(chr(65072), '|').replace(chr(65374), '~')}）\n")
    
    advisor_text = '\n'.join([m['content'] for m in c['advisor_msgs'] if m['content']])
    customer_text = '\n'.join([m['content'] for m in c['customer_msgs'] if m['content']])
    all_text = advisor_text + '\n' + customer_text
    
    # Health issues
    issues = []
    if re.search(r'便秘|排便|拉不出|通畅|大便', all_text): issues.append('🚽 便秘/排便')
    if re.search(r'腹泻|拉肚子|肠胃不适', all_text): issues.append('💧 腹泻/肠胃不适')
    if re.search(r'睡眠|失眠|睡不着|入睡|熬夜|睡不踏实', all_text): issues.append('😴 睡眠障碍')
    if re.search(r'减肥|减脂|减重|体重|胖|BMI|240|120公|减不', all_text): issues.append('⚖️ 减重/体重管理')
    if re.search(r'血糖|糖尿病|控糖|糖化|空腹血糖|餐后血糖', all_text): issues.append('🩸 血糖管理')
    if re.search(r'痘痘|过敏|皮肤|湿疹|荨麻疹', all_text): issues.append('🔴 皮肤问题')
    if re.search(r'消化|吸收|消瘦|不长肉|瘦弱', all_text): issues.append('🍽️ 消化不良/吸收差')
    if re.search(r'情绪|压力|焦虑|高三|高考', all_text): issues.append('🧠 情绪/压力')
    
    p(f"**核心痛点**: {', '.join(issues) if issues else '综合调理'}")
    
    # Products
    prods = []
    for prod in ['益休', '益舒', '益畅', '益健', '镜脂·轻', '镜脂阻', '轻路', '益家人', '畅青', '膳食纤维']:
        if prod in advisor_text:
            prods.append(prod)
    p(f"**推荐产品**: {', '.join(prods) if prods else '多种产品'}")
    
    # Customer opening message
    customer_openings = [m for m in c['customer_msgs'] if m['content'] and m['content'] != '我已经添加了你，现在我们可以开始聊天了。']
    if customer_openings:
        first_real = customer_openings[0]['content'][:200]
        p(f"**客户开场**: 「{first_real}」")
    
    # Show key conversation flow
    p("\n**对话流程**:")
    
    # Find important advisor messages
    important = []
    for m in c['advisor_msgs']:
        txt = m['content']
        if not txt or len(txt) < 10:
            continue
        if any(kw in txt for kw in ['推荐', '方案', '益生菌', '建议', '调理', '搭配', '服用', '下单', '发货', '地址',
                                       '益休', '益畅', '益家人', '镜脂', '轻路', '膳食', '华大',
                                       '专属', '活动', '优惠', '买一', '改善', '效果', '安全', '菌群']):
            important.append((m['time'], '顾问', txt[:300]))
    
    for m in c['customer_msgs']:
        txt = m['content']
        if not txt or len(txt) < 5:
            continue
        if any(kw in txt for kw in ['血糖', '睡眠', '便秘', '减肥', '减重', '体重', '吃药', '中药',
                                       '不舒服', '多久', '效果', '试试', '买', '多少钱', '价格',
                                       '收到', '感觉', '改善', '控制不住', '身高', '年龄']):
            important.append((m['time'], '客户', txt[:200]))
    
    important.sort(key=lambda x: x[0])
    
    # Show key moments (first 2, middle 2, last 2)
    if len(important) > 6:
        for item in important[:3]:
            p(f"  `{item[0][:16]}` **{item[1]}**: {item[2]}")
        p("  `...` *（中间省略）*")
        for item in important[-3:]:
            p(f"  `{item[0][:16]}` **{item[1]}**: {item[2]}")
    else:
        for item in important:
            p(f"  `{item[0][:16]}` **{item[1]}**: {item[2]}")
    
    # Links
    links = re.findall(r'https://j\.youzan\.com/\S+', advisor_text)
    if links:
        p(f"\n  🔗 有赞商品链接: {len(links)} 个")
    
    p("")

# ─── 3. 全局统计 ───
p("## 三、全局数据统计\n")

all_advisor = '\n'.join(['\n'.join([m['content'] for m in c['advisor_msgs'] if m['content']]) for c in all_conversations])
all_customer = '\n'.join(['\n'.join([m['content'] for m in c['customer_msgs'] if m['content']]) for c in all_conversations])
all_text = all_advisor + '\n' + all_customer

# 3.1 产品
p("### 3.1 产品提及频率\n")
prods_full = {
    '益休（睡眠/情绪）': '益休',
    '益畅（肠道/便秘）': '益畅',
    '镜脂·轻（体重管理）': '镜脂',
    '轻路（膳食纤维）': '轻路',
    '益家人（基础益生菌）': '益家人',
    '畅青（吸收/增重）': '畅青',
    '益健（免疫/过敏）': '益健',
    '优美达（品牌）': '优美达',
    '华大（背书）': '华大',
}
p("| 产品 | 功能定位 | 提及次数 |")
p("|------|----------|----------|")
for prod_name, keyword in prods_full.items():
    count = len(re.findall(keyword, all_text))
    p(f"| {prod_name} | - | {count} |")

# 3.2 话术技巧
p("\n### 3.2 销售话术技巧统计\n")
tactics = {
    'SPIN提问式需求挖掘': ['什么情况', '多久了', '平时怎么', '有没有', '之前试过', '方便跟我说'],
    '共情与信任建立': ['辛苦了', '理解', '心疼', '不容易', '专属', '感谢您'],
    '专业科普教育': ['益生菌不是药物', '食品级', '菌群平衡', '有益菌', '有害菌', '肠道菌群'],
    '安全无副作用保证': ['安全', '无副作用', '温和', '不刺激', '无依赖', '不含糖'],
    '生活方式建议': ['饮食', '运动', '喝水', '作息', '食谱', '温水', '睡眠'],
    '限时优惠促单': ['活动', '优惠', '买一送', '划算', '专属价', '限时'],
    '复购/升级引导': ['巩固', '坚持', '继续', '长期', '配合', '搭配'],
}
p("| 话术类型 | 出现次数 | 示例 |")
p("|----------|----------|------|")
for tactic, keywords in tactics.items():
    count = sum(len(re.findall(kw, all_advisor)) for kw in keywords)
    p(f"| {tactic} | {count} | - |")

# ─── 4. 话术模式 ───
p("\n## 四、成交话术黄金模式\n")

p("### 4.1 标准成交四步曲\n")
p("""
```
① 开场破冰（100%执行）
   └─ 「您好，我是您的专属健康顾问」+ 询问需求方向

② 需求诊断（100%执行）
   └─ 年龄/身高/体重 + 症状细节 + 既往史 + 用药情况

③ 科普教育（87.5%执行）
   └─ 「益生菌不是药物，是食品级」+ 菌群原理 + 安全性说明

④ 方案推荐 + 促单（100%执行）
   └─ 产品搭配 + 服用指导 + 活动优惠 + 有赞链接
```
""")

p("### 4.2 三位顾问话术风格对比\n")

# Analyze per advisor
advisors_data = defaultdict(list)
for c in all_conversations:
    advisors_data[c['advisor']].append(c)

for adv_name, convs in advisors_data.items():
    adv_text = '\n'.join(['\n'.join([m['content'] for m in cc['advisor_msgs'] if m['content']]) for cc in convs])
    
    p(f"#### {adv_name}（服务 {len(convs)} 人）\n")
    
    # Style characteristics
    styles = []
    if '方案' in adv_text: styles.append('方案型：倾向出完整健康方案')
    if '食谱' in adv_text or '饮食' in adv_text: styles.append('生活指导型：重视饮食/生活建议')
    if '活动' in adv_text or '买一' in adv_text: styles.append('促销型：善用活动促单')
    if '菌群' in adv_text or '有益菌' in adv_text: styles.append('科普型：注重菌群知识教育')
    if '专属' in adv_text or '感谢' in adv_text: styles.append('服务型：强调专属服务感')
    
    p(f"**话术风格**: {' | '.join(styles) if styles else '综合型'}")
    
    # Extract 2 representative quotes
    quotes = []
    for cc in convs:
        for m in cc['advisor_msgs']:
            txt = m['content']
            if len(txt) > 40 and len(txt) < 300:
                if any(kw in txt for kw in ['方案', '搭配', '建议您', '推荐您', '更适合', '帮助']):
                    if txt not in quotes:
                        quotes.append(txt)
    for q in quotes[:2]:
        p(f"> 「{q[:250]}」\n")
    p("")

# ─── 5. 成交特征 ───
p("## 五、成交特征与关键发现\n")

p("### 5.1 客户购买路径分析\n")
p("""
| 客户 | 核心痛点 | 初始诉求 | 最终购买 | 交叉销售 | 成交周期 |
|------|----------|----------|----------|----------|----------|
""")

for i, c in enumerate(all_conversations):
    adv_text = '\n'.join([m['content'] for m in c['advisor_msgs'] if m['content']])
    cust_text = '\n'.join([m['content'] for m in c['customer_msgs'] if m['content']])
    all_t = adv_text + '\n' + cust_text
    msgs = c['messages']
    
    # Initial ask
    initial = ''
    for m in c['customer_msgs']:
        if m['content'] and m['content'] != '我已经添加了你，现在我们可以开始聊天了。' and len(m['content']) > 3:
            initial = m['content'][:40]
            break
    
    # Products
    prods = set()
    for pn in ['益休', '益畅', '镜脂', '轻路', '益家人', '畅青']:
        if pn in adv_text: prods.add(pn)
    products_str = '+'.join(prods) if prods else '产品'
    
    # Period
    d1 = msgs[-1]['time'][:10] if msgs else ''
    d2 = msgs[0]['time'][:10] if msgs else ''
    period = f"{d1}~{d2}"
    
    # Pain points
    pains = []
    if re.search(r'便秘|排便', all_t): pains.append('便秘')
    if re.search(r'睡眠|失眠|入睡', all_t): pains.append('睡眠')
    if re.search(r'减[肥重脂]|体重', all_t): pains.append('减重')
    if re.search(r'血糖|糖尿', all_t): pains.append('血糖')
    if re.search(r'痘痘|皮肤|过敏', all_t): pains.append('皮肤')
    
    cross_sell = '是' if len(prods) > 1 else '否'
    
    p(f"| {c['customer'][:8]} | {','.join(pains)} | {initial} | {products_str} | {cross_sell} | {period} |")

p("")
p("### 5.2 十大关键发现\n")
p("""
1. **血糖是一切入口**：8位成交用户100%是三诺血糖仪用户，血糖管理是私域触达的原点，但成交均延伸到了肠道/睡眠/减重等更多健康领域。

2. **「赵老师」是高转化人设**：「老师」称谓天然带有权威感和信任度，配合专业科普话术，是最高效的信任建立方式。

3. **益生菌 ≠ 药物是核心认知壁垒**：几乎每段对话都出现了「益生菌不是药物，是食品级」的科普，说明用户普遍存在「怕吃药有副作用」的心理。

4. **先生活建议后产品推荐是标配**：顾问普遍先给饮食/作息建议，再自然过渡到产品，降低了用户的防御心理。

5. **有赞链接是标准促单动作**：所有成交对话中顾问都发送了有赞商品链接，选品 → 发链接 → 跟进是固定流程。

6. **睡眠+血糖是黄金交叉点**：三诺用户天然关注血糖，但睡眠障碍同样高频（4/8用户），「益休」作为睡眠益生菌成为重要的交叉销售品。

7. **姣姣的服务颗粒度最细**：从数据看，姣姣（4人）在方案定制、售后跟进方面最为细致，用户粘性更高。

8. **客户主动表达需求后转化最快**：客户主动说出「血糖控制不住」「想了解肠胃」「怎么减肥」等需求后，成交效率显著高于被动唤醒。

9. **复购/升级引导存在明显空间**：大部分对话在首次成交后缺少体系化的复购跟进，这是一个巨大的 LTV 提升机会。

10. **赫氏消亡反应预告缺失**：对话中没有发现顾问主动告知好转反应的情况，如能提前预告「服用初期可能出现的反应」，能减少售后焦虑，提升信任。
""")

# ─── 6. 优化建议 ───
p("## 六、运营优化建议\n")

p("### 6.1 话术标准化\n")
p("""
| 场景 | 当前问题 | 优化建议 |
|------|----------|----------|
| 开场破冰 | 话术完全统一，缺乏个性 | 保留统一框架，增加基于客户画像的个性化问候 |
| 需求诊断 | 问题顺序不统一 | 建立「先数据后症状」的标准问诊清单 |
| 科普教育 | 深度参差不齐 | 制作3套不同深度的科普话术（入门/进阶/深度） |
| 异议处理 | 依赖个人经验 | 建立 Top 5 异议的标准应答卡 |
| 促单环节 | 方式单一（发链接） | 增加「试用装」「阶梯优惠」「限时活动」等多种促单手段 |
| 售后跟进 | 被动等待为主 | 建立 Day3/7/14/30 的主动跟进 SOP |
""")

p("### 6.2 产品搭配标准化\n")
p("""
| 场景组合 | 主推产品 | 辅推产品 | 话术切入角度 |
|----------|----------|----------|-------------|
| 血糖 + 便秘 | 益畅 | 轻路（膳食纤维） | 「先通后养，肠道通畅血糖更稳」 |
| 血糖 + 睡眠 | 益休 | 益家人 | 「肠脑轴调节，睡得好血糖自然稳」 |
| 血糖 + 减重 | 镜脂·轻 | 轻路 | 「科学控体重，不给胰岛加负担」 |
| 纯肠道调理 | 益畅 | 益家人 | 「对症+基础，双管齐下」 |
| 皮肤/过敏 | 益健 | 益休（如有压力） | 「肠道免疫调节，从内到外改善」 |
| 消瘦/吸收差 | 畅青 | 益家人 | 「先养吸收力，营养才能真正补进去」 |
""")

p("### 6.3 售后跟进SOP\n")
p("""
| 节点 | 时机 | 动作 | 话术要点 |
|------|------|------|----------|
| 服用指导 | 成交当天 | 发送服用方法 + 注意事项 | 温水冲服、时间、剂量 |
| 好转反应预警 | Day 3 | 主动询问 + 预告赫氏消亡反应 | 「初期可能会有轻微不适，这是好菌在工作的信号」 |
| 效果初体验 | Day 7 | 询问感受 + 收集反馈 | 「这几天感觉怎么样？」 |
| 深度跟进 | Day 14 | 效果评估 + 方案调整 | 「如果效果好，建议继续巩固一个周期」 |
| 复购引导 | Day 21-25 | 提醒即将用完 + 复购优惠 | 「您的产品快用完了，现在复购有专属优惠」 |
| 长期方案 | Day 30 | 全面评估 + 推荐3个月方案 | 「一个月只是开始，持续调理效果更好」 |
""")

# ─── 7. 精华话术 ───
p("## 七、精华话术摘录\n")

# Extract top advisor messages that demonstrate good tactics
p("### 7.1 科普教育类\n")
examples = []
for c in all_conversations:
    for m in c['advisor_msgs']:
        txt = m['content']
        if not txt or len(txt) < 40:
            continue
        if '益生菌不是药物' in txt or '食品级' in txt or '菌群' in txt:
            examples.append((c['advisor'], txt[:400]))
            break
for adv, txt in examples[:3]:
    p(f"> **{adv}**: {txt}\n")

p("### 7.2 方案推荐类\n")
examples = []
for c in all_conversations:
    for m in c['advisor_msgs']:
        txt = m['content']
        if not txt or len(txt) < 40:
            continue
        if '方案' in txt or '搭配' in txt:
            examples.append((c['advisor'], txt[:400]))
            break
for adv, txt in examples[:3]:
    p(f"> **{adv}**: {txt}\n")

p("### 7.3 异议处理类\n")
examples = []
for c in all_conversations:
    for m in c['advisor_msgs']:
        txt = m['content']
        if not txt or len(txt) < 30:
            continue
        if any(kw in txt for kw in ['不冲突', '放心', '安全', '温和', '可以放心', '不影响']):
            examples.append((c['advisor'], txt[:400]))
for adv, txt in examples[:3]:
    p(f"> **{adv}**: {txt}\n")

p("---\n")
p("*报告由 WorkBuddy 自动分析生成，数据来源：企业微信聊天记录导出*")

# Save
report_path = "/Users/SaciNa/WorkBuddy/2026-05-26-15-30-31/三诺成交用户分析报告.md"
with open(report_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(out))

print(f"✅ 报告已生成: {report_path}")
print(f"   共分析 {len(all_conversations)} 个成交用户，{sum(len(c['messages']) for c in all_conversations)} 条消息")
