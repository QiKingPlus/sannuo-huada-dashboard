#!/usr/bin/env python3
"""更新三诺华大成交数据分析和报告"""
import openpyxl, json, os
from datetime import datetime
from collections import Counter, defaultdict

WORK_DIR = "/Users/SaciNa/WorkBuddy/2026-05-26-15-30-31"
ORDER_PATH = "/Users/SaciNa/Documents/臻选&芥子/臻选私域/三诺&华大成交用户清单.xlsx"

# ============================================================
# Part 1: 读取所有订单数据
# ============================================================
wb = openpyxl.load_workbook(ORDER_PATH)
ws = wb.active
headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]

orders = []
for row in range(2, ws.max_row + 1):
    o = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row, col).value
        o[headers[col-1]] = str(val) if val is not None else ""
    orders.append(o)
wb.close()

# 数据清洗
for o in orders:
    # 清理下单时间
    t = o.get("下单时间", "")
    t = t.replace("下单时间：", "").strip()
    o["clean_time"] = t
    # 清理产品名
    pname = o.get("产品名称", "")
    if "益休" in pname: o["product_short"] = "益休"
    elif "镜脂" in pname: o["product_short"] = "镜脂套组"
    elif "轻路" in pname: o["product_short"] = "轻路"
    elif "畅青" in pname: o["product_short"] = "畅青"
    elif "益家人" in pname: o["product_short"] = "益家人"
    elif "益畅" in pname: o["product_short"] = "益畅"
    else: o["product_short"] = "其他"
    # 价格
    o["price_float"] = float(o.get("产品价格", 0))
    # 件数
    o["qty"] = int(o.get("件数", "1").replace("件", "")) if o.get("件数") else 1
    # 年龄
    try: o["age_int"] = int(o.get("年龄", 0))
    except: o["age_int"] = 0

total_orders = len(orders)
unique_phones = set(o.get("手机", "") for o in orders)
total_users = len(unique_phones)
total_revenue = sum(o["price_float"] for o in orders)

# ============================================================
# Part 2: 销售人效
# ============================================================
sales_data = defaultdict(lambda: {"orders": 0, "revenue": 0, "users": set(), "products": Counter()})
for o in orders:
    s = o.get("销售", "未知")
    sales_data[s]["orders"] += 1
    sales_data[s]["revenue"] += o["price_float"]
    sales_data[s]["users"].add(o.get("手机", ""))
    sales_data[s]["products"][o["product_short"]] += o["price_float"]

# ============================================================
# Part 3: 产品营收
# ============================================================
product_data = defaultdict(lambda: {"qty": 0, "revenue": 0, "count": 0})
for o in orders:
    p = o["product_short"]
    product_data[p]["qty"] += o["qty"]
    product_data[p]["revenue"] += o["price_float"]
    product_data[p]["count"] += 1

# ============================================================
# Part 4: 时间分布
# ============================================================
time_dist = Counter()
date_dist = Counter()
for o in orders:
    t = o["clean_time"]
    if t:
        try:
            dt = datetime.strptime(t, "%Y-%m-%d %H:%M:%S")
            hour = dt.hour
            date = dt.strftime("%m-%d")
            if 6 <= hour < 9: slot = "06-09点"
            elif 9 <= hour < 12: slot = "09-12点"
            elif 12 <= hour < 14: slot = "12-14点"
            elif 14 <= hour < 18: slot = "14-18点"
            elif 18 <= hour < 21: slot = "18-21点"
            else: slot = "21-06点"
            time_dist[slot] += 1
            date_dist[date] += 1
        except:
            pass

# ============================================================
# Part 5: 客群画像
# ============================================================
ages = [o["age_int"] for o in orders if o["age_int"] > 0]
genders = Counter(o.get("性别", "") for o in orders)
provinces = Counter(o.get("省份", "").strip() for o in orders)
demands = Counter(o.get("需求", "") for o in orders)

age_groups = {"30-39岁": 0, "40-49岁": 0, "50-59岁": 0, "60岁以上": 0}
for a in ages:
    if a < 40: age_groups["30-39岁"] += 1
    elif a < 50: age_groups["40-49岁"] += 1
    elif a < 60: age_groups["50-59岁"] += 1
    else: age_groups["60岁以上"] += 1

# ============================================================
# Part 6: 复购分析
# ============================================================
phone_orders = defaultdict(list)
for o in orders:
    phone_orders[o.get("手机", "")].append(o)
repeat_users = {p: os for p, os in phone_orders.items() if len(os) > 1}

# ============================================================
# Part 7: 生成 Markdown 报告
# ============================================================
lines = []
def add(s=""): lines.append(s)

add("# 三诺&华大成交订单分析报告（更新版）")
add("")
add(f"> **数据更新**: {datetime.now().strftime('%Y-%m-%d %H:%M')} | **订单总数**: {total_orders} 条 | **去重用户**: {total_users} 人 | **总营收**: ¥{total_revenue:,.2f}")
add(f"> **对比上版**: 订单 11→{total_orders} (+4), 用户 9→{total_users} (+3), 营收 ¥4,307.80→¥{total_revenue:,.2f} (+¥{total_revenue-4307.80:,.2f})")
add("")

# ---- 营收总览 ----
add("## 一、营收总览")
add("")
add("| 指标 | 上版(11单) | 更新(15单) | 变化 |")
add("|------|-----------|-----------|------|")
add(f"| 统计周期 | 05-21~05-26 | 05-21~05-27 | +1天 |")
add(f"| 总订单数 | 11 | {total_orders} | +4 (+36%) |")
add(f"| 去重用户 | 9 | {total_users} | +3 (+33%) |")
add(f"| 总营收 | ¥4,307.80 | ¥{total_revenue:,.2f} | +¥{total_revenue-4307.80:,.2f} (+{(total_revenue/4307.80-1)*100:.0f}%) |")
add(f"| 平均客单价 | ¥391.62 | ¥{total_revenue/total_orders:,.2f} | ¥{total_revenue/total_orders-391.62:+,.2f} |")
add(f"| 用户ARPU | ¥538.48 | ¥{total_revenue/total_users:,.2f} | ¥{total_revenue/total_users-538.48:+,.2f} |")
add(f"| 复购用户 | 2人(25%) | {len(repeat_users)}人({len(repeat_users)/total_users*100:.0f}%) | — |")
add("")

# ---- 新增订单 ----
add("## 二、新增订单明细（4单）")
add("")
add("| 销售 | 收件人 | 产品 | 价格 | 件数 | 年龄 | 性别 | 需求 | 地区 |")
add("|------|--------|------|------|------|------|------|------|------|")

# 旧手机号集合（前11单）
old_set = {"18190919791","13501134294","13986460439","18967469708","17703100088","15262003719","13975315486","15073313688","15911579927"}
for o in orders:
    phone = o.get("手机", "")
    if phone not in old_set:
        add(f"| {o.get('销售','')} | {o.get('收件人','')} | {o['product_short']} | ¥{o['price_float']:,.2f} | {o['qty']}件 | {o.get('年龄','')} | {o.get('性别','')} | {o.get('需求','')} | {o.get('省份','')}{o.get('城市','')} |")

add("")
add("> **王庆丰** 镜脂套组×3件 ¥1,699.80 为目前最高客单价订单，来自赵老师，62岁内蒙古用户。")

# ---- 销售人效 ----
add("")
add("## 三、销售人效排行")
add("")
add("| 排名 | 销售 | 订单数 | 营收 | 服务用户 | 均单值 | 人均ARPU | 主力产品 |")
add("|------|------|--------|------|---------|--------|---------|---------|")
s_sorted = sorted(sales_data.items(), key=lambda x: x[1]["revenue"], reverse=True)
for i, (name, data) in enumerate(s_sorted, 1):
    top_p = data["products"].most_common(2)
    top_products = " + ".join(f"{p}(¥{v:,.0f})" for p, v in top_p)
    add(f"| {i} | **{name}** | {data['orders']} | ¥{data['revenue']:,.2f} | {len(data['users'])}人 | ¥{data['revenue']/data['orders']:,.2f} | ¥{data['revenue']/len(data['users']):,.2f} | {top_products} |")

add("")

# 人效对比
wu = sales_data.get("吴老师", {})
zhao = sales_data.get("赵老师", {})
wang = sales_data.get("王老师", {})

add(f"**关键对比**:")
add(f"- 赵老师 {zhao.get('orders',0)}单 ¥{zhao.get('revenue',0):,.2f} vs 吴老师 {wu.get('orders',0)}单 ¥{wu.get('revenue',0):,.2f} — 赵老师营收领先")
if zhao.get('revenue',0) > 0 and wang.get('revenue',0) > 0:
    add(f"- 赵老师营收是王老师的 **{zhao['revenue']/wang['revenue']:.1f}x**")

# ---- 产品营收 ----
add("")
add("## 四、产品营收排行")
add("")
add("| 排名 | 产品 | 功能 | 销量(件) | 订单数 | 营收 | 占比 | 均价 |")
add("|------|------|------|---------|--------|------|------|------|")
p_sorted = sorted(product_data.items(), key=lambda x: x[1]["revenue"], reverse=True)
product_func = {"益休":"睡眠","镜脂套组":"轻体管理","轻路":"膳食纤维","益家人":"基础益生菌","畅青":"吸收/增重","益畅":"便秘"}
for i, (p, data) in enumerate(p_sorted, 1):
    func = product_func.get(p, "")
    pct = data['revenue'] / total_revenue * 100
    avg = data['revenue'] / data['count'] if data['count'] > 0 else 0
    medal = ["🥇","🥈","🥉","4","5","6"][i-1]
    add(f"| {medal} | {p} | {func} | {data['qty']} | {data['count']} | ¥{data['revenue']:,.2f} | {pct:.1f}% | ¥{avg:,.2f} |")

add("")

# 集中度分析
top2_rev = sum(d['revenue'] for p, d in p_sorted[:2])
add(f"**产品集中度**: 前2产品贡献 {top2_rev/total_revenue*100:.0f}% 营收。")

# ---- 客群画像 ----
add("")
add("## 五、客群画像")
add("")
add(f"- **年龄范围**: {min(ages)}~{max(ages)} 岁，平均 **{sum(ages)/len(ages):.0f}** 岁")
add(f"- **年龄分布**: " + " | ".join(f"{k} {v}单" for k, v in age_groups.items()))
add(f"- **性别**: 男 {genders.get('男',0)}单 / 女 {genders.get('女',0)}单")
add(f"- **Top5 省份**: " + " > ".join(f"{p}({c}单)" for p, c in provinces.most_common(5)))
add(f"- **需求分布**: " + " > ".join(f"{d}({c}单)" for d, c in demands.most_common(6)))
add("")

# ---- 时间分布 ----
add("## 六、下单时间分布")
add("")
add("| 时段 | 订单数 | 占比 |")
add("|------|--------|------|")
for slot in ["06-09点","09-12点","12-14点","14-18点","18-21点","21-06点"]:
    c = time_dist.get(slot, 0)
    add(f"| {slot} | {c} | {c/total_orders*100:.0f}% |")
add("")
add(f"**高峰日**: " + " > ".join(f"{d}({c}单)" for d, c in date_dist.most_common(5)))

# ---- 复购行为 ----
add("")
add("## 七、复购行为")
add("")
if repeat_users:
    for phone, os_list in repeat_users.items():
        user_name = os_list[0].get("收件人", phone)
        sales_name = os_list[0].get("销售", "")
        products = " → ".join(o["product_short"] for o in os_list)
        total = sum(o["price_float"] for o in os_list)
        add(f"- **{user_name}** ({sales_name}): {products} | 合计 ¥{total:,.2f} | {len(os_list)}单")
else:
    add("暂无复购用户")

add("")
add("## 八、综合洞察与建议")
add("")
add("### 8.1 本轮数据变化亮点")
add(f"1. **营收增长 {(total_revenue/4307.80-1)*100:.0f}%**: 7天内从 ¥4,308 增至 ¥{total_revenue:,.2f}")
add(f"2. **客单价突破**: 王庆丰镜脂套组×3件 ¥1,699.80 创新高，验证了高客单价产品的批量转化可能")
add("3. **地域扩展**: 新增内蒙古（王庆丰）、广东（陈志），不再局限于此前集中的北京/四川")
add("4. **姣姣持续高产**: 吴老师累计8单，稳居单量第一，且新成交了镜脂套组高客单价产品")
add("5. **冯琳案例值得深挖**: 36岁妊娠糖尿病患者，从血糖管理切入→发现睡眠问题→成交益休，路径非常典型")
add("")
add("### 8.2 冯琳对话亮点（新增聊天记录123条）")
add("- **对话周期**: 3天(05-25~05-27)，用户52条/顾问71条，互动率极高")
add("- **核心路径**: 用户自述睡眠问题 → 姣姣问诊(年龄/身高/体重/血糖/用药) → 发现多梦浅睡1-2年+褪黑素无效 → 引导脑肠轴机制 → 推荐益休 → 用户确认「多久见效」→ 成交")
add("- **值得复用的手法**: ①先解决饮食困惑(杂粮饭胃疼→推荐替代方案)，建立专业信任 ②问诊围绕「血糖+睡眠」双线并进 ③用褪黑素无效引出益生菌差异化价值 ④持续3天跟进血糖数据，关怀感强")
add("- **SOP优化印证**: 本次对话完全符合我们SOP的「先生活建议→再产品推荐」节奏，且问诊深度够，成交自然")
add("")
add("### 8.3 后续关注")
add("1. 冯琳益休效果反馈(预计7-14天)，可做案例素材")
add("2. 王庆丰大单复购可能性(镜脂套组×3=约3个月用量)")
add("3. 陈志(广州)和施俊威(金华)有无对应聊天记录？可补充分析")
add("")

# ============================================================
# 保存报告
# ============================================================
report_path = os.path.join(WORK_DIR, "三诺华大成交订单分析报告_更新版.md")
with open(report_path, "w", encoding="utf-8") as f:
    f.write("\n".join(lines))

print(f"报告已保存: {report_path}")
print(f"总订单: {total_orders}, 总用户: {total_users}, 总营收: ¥{total_revenue:,.2f}")
print(f"新增4单: ¥{total_revenue-4307.80:,.2f}")
